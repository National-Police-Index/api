import pandas as pd
from openpyxl import load_workbook

def filter_unmatched_with_candidates(input_file: str, output_file: str):
    """
    Read unmatched_debug.xlsx and create a new file with only sheets 
    that have candidate matches (excluding 'NO CANDIDATES FOUND' sheets)
    """
    # Load the workbook
    wb = load_workbook(input_file)
    
    # Create a new Excel writer for the filtered output
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        sheets_with_candidates = 0
        sheets_without_candidates = 0
        
        # Always copy the Summary sheet first if it exists
        if 'Summary' in wb.sheetnames:
            summary_df = pd.read_excel(input_file, sheet_name='Summary')
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Process each sheet
        for sheet_name in wb.sheetnames:
            if sheet_name == 'Summary':
                continue
                
            # Read the sheet
            df = pd.read_excel(input_file, sheet_name=sheet_name)
            
            # Check if the sheet contains "NO CANDIDATES FOUND"
            # This will be in the dataframe as a value
            has_no_candidates = df.astype(str).apply(
                lambda row: row.str.contains('NO CANDIDATES FOUND', case=False).any(), 
                axis=1
            ).any()
            
            if not has_no_candidates:
                # This sheet has candidates, so include it
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                sheets_with_candidates += 1
                print(f"✓ Included: {sheet_name}")
            else:
                sheets_without_candidates += 1
                print(f"✗ Excluded: {sheet_name} (no candidates)")
        
        print(f"\nSummary:")
        print(f"Sheets with candidates: {sheets_with_candidates}")
        print(f"Sheets without candidates: {sheets_without_candidates}")
        print(f"Output saved to: {output_file}")


if __name__ == "__main__":
    input_file = "data/output/unmatched_debug.xlsx"
    output_file = "data/output/unmatched_with_candidates.xlsx"
    
    filter_unmatched_with_candidates(input_file, output_file)