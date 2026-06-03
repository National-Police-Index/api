"""
Convert matched_clean_with_conflicts.xlsx to JSON format for front-end validation UI.

This script restructures the Excel output into a JSON format that makes it easy for
users to validate officer matches by comparing:
1. The matched officer's details and employment history
2. All other officers with the same name in the database
3. The source document where the officer was mentioned
"""

import pandas as pd
import json
from typing import List, Dict, Any
import openpyxl
from pathlib import Path


def parse_officer_info_section(sheet) -> Dict[str, Any]:
    """
    Parse the officer info header section from an Excel sheet.
    
    Returns a dict with keys like:
    - mention_uid
    - provisional_case_name
    - document_link
    - matched_post_id
    - match_probability
    - first_name, middle_name, last_name
    - matched_agency
    - mentioned_agencies
    - total_employment_stints
    - other_officers_count
    """
    officer_info = {}
    
    # Read the first ~13 rows which contain officer info in Field/Value pairs
    for row in sheet.iter_rows(min_row=2, max_row=14, values_only=True):
        if row[0] and row[1]:  # Field and Value columns
            field = str(row[0]).strip()
            value = row[1]
            
            # Map field names to JSON keys
            field_mapping = {
                "Mention UID": "mention_uid",
                "Provisional Case Name": "provisional_case_name",
                "Document Link": "document_link",
                "Matched POST Person Number": "matched_post_id",
                "Match Probability": "match_probability",
                "First Name": "first_name",
                "Middle Name": "middle_name",
                "Last Name": "last_name",
                "Matched Agency": "matched_agency",
                "Mentioned Agencies": "mentioned_agencies",
                "Total Employment Stints (Matched Officer)": "total_employment_stints",
                "Other Officers with Same Name in Database": "other_officers_summary"
            }
            
            if field in field_mapping:
                officer_info[field_mapping[field]] = value
    
    return officer_info


def parse_other_officers_section(sheet, start_row: int) -> tuple[List[Dict[str, Any]], int]:
    """
    Parse the "OTHER OFFICERS NAMED..." section.
    
    Returns: (list of other officer records, next_row_to_read)
    """
    other_officers = []
    
    # Find the header row (contains post_person_nbr, post_first_name, etc.)
    header_row = None
    for i, row in enumerate(sheet.iter_rows(min_row=start_row, max_row=start_row + 10, values_only=True), start=start_row):
        if row and 'post_person_nbr' in str(row[0]).lower():
            header_row = i
            break
    
    if not header_row:
        return [], start_row
    
    # Get column names
    headers = [cell for cell in sheet[header_row] if cell.value]
    header_names = [h.value for h in headers]
    
    # Read data rows until we hit an empty row or the next section
    current_row = header_row + 1
    while current_row <= sheet.max_row:
        row_values = [cell.value for cell in sheet[current_row]]
        
        # Stop if empty row or next section header
        if not any(row_values) or (row_values[0] and 'FULL EMPLOYMENT HISTORY' in str(row_values[0])):
            break
        
        # Create record dict
        if row_values[0]:  # Has post_person_nbr
            record = {}
            for i, header in enumerate(header_names):
                if i < len(row_values):
                    record[header] = row_values[i]
            other_officers.append(record)
        
        current_row += 1
    
    return other_officers, current_row


def parse_matched_officer_history(sheet, start_row: int) -> List[Dict[str, Any]]:
    """
    Parse the "FULL EMPLOYMENT HISTORY FOR MATCHED OFFICER" section.
    
    Returns: list of employment stint records
    """
    employment_history = []
    
    # Find the header row
    header_row = None
    for i in range(start_row, min(start_row + 10, sheet.max_row + 1)):
        row_values = [cell.value for cell in sheet[i]]
        if row_values and 'post_person_nbr' in str(row_values[0]).lower():
            header_row = i
            break
    
    if not header_row:
        return []
    
    # Get column names
    headers = [cell for cell in sheet[header_row] if cell.value]
    header_names = [h.value for h in headers]
    
    # Read data rows
    current_row = header_row + 1
    while current_row <= sheet.max_row:
        row_values = [cell.value for cell in sheet[current_row]]
        
        # Stop if empty row
        if not any(row_values):
            break
        
        # Create record dict
        if row_values[0]:  # Has post_person_nbr
            record = {}
            for i, header in enumerate(header_names):
                if i < len(row_values):
                    record[header] = row_values[i]
            employment_history.append(record)
        
        current_row += 1
    
    return employment_history


def convert_excel_to_json(excel_path: str, output_path: str = None) -> List[Dict[str, Any]]:
    """
    Convert matched_clean_with_conflicts.xlsx to JSON format.
    
    Output JSON structure:
    [
        {
            "officer_info": {
                "mention_uid": "...",
                "provisional_case_name": "...",
                "document_link": "...",
                "matched_post_id": "...",
                "match_probability": 0.8052,
                "first_name": "MICHAEL",
                "middle_name": "JAMES",
                "last_name": "WEBB",
                "matched_agency": "SAN DIEGO POLICE DEPARTMENT",
                "mentioned_agencies": "None",
                "total_employment_stints": 4,
                "other_officers_summary": "10 unique officer(s), 16 total record(s)"
            },
            "other_officers_with_same_name": [
                {
                    "post_person_nbr": "A25-D11",
                    "post_first_name": "MICHAEL",
                    "post_middle_name": "D",
                    "post_last_name": "WEBB",
                    "post_agency_name": "ANAHEIM POLICE DEPARTMENT",
                    "post_start_date": "...",
                    "post_end_date": "..."
                },
                ...
            ],
            "matched_officer_employment_history": [
                {
                    "post_person_nbr": "B05-C48",
                    "post_first_name": "MICHAEL",
                    "post_middle_name": "JAMES",
                    "post_last_name": "WEBB",
                    "post_agency_name": "SAN DIEGO POLICE DEPARTMENT",
                    "post_start_date": "...",
                    "post_end_date": "...",
                    "separation_reason": "Retired"
                },
                ...
            ],
            "validation": {
                "status": "pending",  // pending | correct | incorrect
                "validated_by": null,
                "validated_at": null,
                "notes": null
            }
        },
        ...
    ]
    """
    workbook = openpyxl.load_workbook(excel_path)
    all_officers = []
    
    # Process each sheet (each sheet = one officer needing validation)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        print(f"Processing sheet: {sheet_name}")
        
        try:
            # Parse officer info header
            officer_info = parse_officer_info_section(sheet)
            
            # Find where "OTHER OFFICERS" section starts (around row 18-20)
            other_officers_start = None
            for i in range(15, min(30, sheet.max_row + 1)):
                cell_value = sheet.cell(row=i, column=1).value
                if cell_value and 'OTHER OFFICERS' in str(cell_value):
                    other_officers_start = i + 2  # Skip header row
                    break
            
            # Parse other officers section
            other_officers = []
            matched_history_start = other_officers_start
            if other_officers_start:
                other_officers, matched_history_start = parse_other_officers_section(
                    sheet, other_officers_start
                )
            
            # Parse matched officer employment history
            matched_history = []
            if matched_history_start:
                # Skip ahead to find "FULL EMPLOYMENT HISTORY" header
                for i in range(matched_history_start, min(matched_history_start + 10, sheet.max_row + 1)):
                    cell_value = sheet.cell(row=i, column=1).value
                    if cell_value and 'FULL EMPLOYMENT HISTORY' in str(cell_value):
                        matched_history = parse_matched_officer_history(sheet, i + 2)
                        break
            
            # Construct the officer validation record
            officer_record = {
                "officer_info": officer_info,
                "other_officers_with_same_name": other_officers,
                "matched_officer_employment_history": matched_history,
                "validation": {
                    "status": "pending",
                    "validated_by": None,
                    "validated_at": None,
                    "notes": None
                }
            }
            
            all_officers.append(officer_record)
            
        except Exception as e:
            print(f"Error processing sheet {sheet_name}: {e}")
            continue
    
    # Write to JSON file
    if output_path:
        with open(output_path, 'w') as f:
            json.dump(all_officers, f, indent=2, default=str)
        print(f"Wrote {len(all_officers)} officer records to {output_path}")
    
    return all_officers


def convert_excel_to_jsonl(excel_path: str, output_path: str = None) -> List[Dict[str, Any]]:
    """
    Convert matched_clean_with_conflicts.xlsx to JSONL format (one object per line).
    
    Better for streaming/incremental loading in front-ends.
    """
    officers = convert_excel_to_json(excel_path, output_path=None)
    
    if output_path:
        with open(output_path, 'w') as f:
            for officer in officers:
                f.write(json.dumps(officer, default=str) + '\n')
        print(f"Wrote {len(officers)} officer records to {output_path}")
    
    return officers


if __name__ == "__main__":
    input_path = "../data/output/interface/matched_clean_with_conflicts.xlsx"

    output_path = "../data/output/interface/conflicts_for_validation.json"
    
    convert_excel_to_jsonl(excel_path=input_path, output_path=output_path)
    
    print("Done!")